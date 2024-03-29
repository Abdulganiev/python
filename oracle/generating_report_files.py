import shutil, zipfile, os, jaydebeapi, json, re, time, stat
import pandas as pd
from pandas.io.excel import ExcelWriter
from datetime import datetime
import datetime as dt

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font

import smtplib # Импортируем библиотеку по работе с SMTP
# Добавляем необходимые подклассы - MIME-типы
import mimetypes # Импорт класса для обработки неизвестных MIME-типов, базирующихся на расширении файла
from email import encoders # Импортируем энкодер
from email.mime.base import MIMEBase # Общий тип
from email.mime.text import MIMEText # Текст/HTML
from email.mime.image import MIMEImage # Изображения
from email.mime.audio import MIMEAudio # Аудио
from email.mime.multipart import MIMEMultipart # Многокомпонентный объект

from sys import platform

# *****************************************************************
mail = 'IVAbdulganiev@yanao.ru'
mail_rezerv = '300195@mail.ru, bers@yandex.ru'
today = dt.date.today()
sec = time.strftime("%S", time.localtime())

# *****************************************************************
if platform == 'linux' or platform == 'linux2': # linux
    trek = '/home/user/schedule'
    server_105 = '/home/share/105'
    server_68 = '/home/share22/s68'
    milk = f'{server_68}/! Обмен с организациями/Деп_Здрав/от них молочка/'
    prezent = f'{server_68}/! Обмен с организациями/Деп_Здрав/от них подарок/'
    samohod = f'{server_68}/! Обмен с организациями/Служба самоход/от них реестр/'
    path_to_gasu = f'{server_68}/! Обмен с организациями/ГМУ/'
    pfr_4454 = f'{server_68}/! Обмен с организациями/от ПФР/prf_4454/'
    vipnet_main = f'{server_68}/! Обмен с организациями/УСЗН/к ним/'
    vipnet_test = f'{trek}/VipoNet_out1/'
    vipnet_main_cso = f'{server_68}/! Обмен с организациями/ЦСО_отправка/'
    vipnet_test_cso = f'{trek}/VipoNet_out1/'
    vipnet_main_do = f'{server_68}/! Обмен с организациями/Деп_Обр/ЭС/'
    vipnet_test_do = f'{trek}/VipoNet_out1/'
elif platform == 'darwin': # OS X
    pass
elif platform == 'win32': # Windows...
    trek = 'd:/python/schedule'
    server_105 = 'Y:'
    server_68 = 'Z:'
    milk = f'{server_68}/! Обмен с организациями/Деп_Здрав/от них молочка/'
    prezent = f'{server_68}/! Обмен с организациями/Деп_Здрав/от них подарок/'
    samohod = f'{server_68}/! Обмен с организациями/Служба самоход/от них реестр/'
    path_to_gasu = f'{server_68}/! Обмен с организациями/ГМУ/'
    pfr_4454 = f'{server_68}/! Обмен с организациями/от ПФР/prf_4454/'
    vipnet_main = 'c:/VipoNet_out/'
    vipnet_test = 'c:/VipoNet_out1/'
    vipnet_main_cso = f'{server_68}/! Обмен с организациями/ЦСО_отправка/'
    vipnet_test_cso = 'c:/VipoNet_out1/'
    vipnet_main_do = f'{server_68}/! Обмен с организациями/Деп_Обр/ЭС/'
    vipnet_test_do = 'c:/VipoNet_out1/'

# *****************************************************************
def get_platform():
    answer = {
        'platform' : platform,
        'trek' : trek,
        'vipnet_main' : vipnet_main,
        'vipnet_test' : vipnet_test,
        'path_to_gasu' : path_to_gasu,
        'server_105' : server_105,
        'server_68' : server_68,
        'milk' : milk,
        'prezent' : prezent,
        'samohod' : samohod,
        'pfr_4454' : pfr_4454,
    }
    return answer

# *****************************************************************
def goto_folder(x=trek):
    os.chdir(x)

# *****************************************************************
def report_1gmu(df, file_name, mail, name_log, test, path_backup):
    path_to = path_to_gasu
    path_in = f'{trek}/GMU/'

    data = pd.DataFrame(df)
    data.to_excel(f'{path_in}{file_name}', index=False)
    
    try:
        os.remove(f'{path_to}{file_name}')
        text = f'{file_name} в папке {path_to} удален'
        writing_to_log_file(name_log, text)
    except:
        pass

    try:
        shutil.move(f'{path_in}{file_name}', path_to)
        text = f'{path_in}{file_name} в папке {path_to}'
        writing_to_log_file(name_log, text)
        text_mail = f'{file_name} в {path_to}'
        text = f'1-ГМУ {file_name}'
        send_email(mail, text, msg_text=text_mail)
    except Exception as e:
        text = f'Ошибка переноса файла {path_in}{file_name} в папку {path_to} - {e}'
        writing_to_log_file(name_log, text)      

# *****************************************************************
def report_death(df, region_id, id):
    name_log = 'checking_death'
    text = f'0{region_id} - Данные по умершему в МО-{region_id} док-{id}'
    file_name = text + '.xlsx'

    data = pd.DataFrame(df)
    data.to_excel(file_name, index=False)

    send_email(mail, text, msg_text=file_name, files=[file_name])

    new_file_name = f'{today} - {file_name}'
    os.replace(file_name, f'{trek}/backup/{new_file_name}') 

    writing_to_log_file(name_log, text)      

# ********************************************************
def alarm_log(mail, log, text):
    writing_to_log_file(log, text)
    send_email(mail, log, msg_text=text)

# *****************************************************************
def generating_list_GKV_kv(name_log, text, file_name, test, mail):
    new_file_name = f'{today} - {file_name}'
    text = f'mail - {mail} \n{text}'

    send_email(mail, file_name, msg_text=text, files=[file_name])
    os.replace(file_name, f'{trek}/backup/{new_file_name}') 
    writing_to_log_file(name_log, text)

# *****************************************************************
def generating_report_GKV_kv(df, name_log, name, region_id, mail):
    data = pd.DataFrame(df)
    file_name = f'{name}.xlsx'
    text = f'{name} на {today}'
    new_file_name = f'{today} - {file_name}'

    if region_id == '58':
        with ExcelWriter(file_name) as writer:
            data.to_excel(writer, sheet_name='104', index=False)
            data.to_excel(writer, sheet_name=f'{region_id}', index=False)
    else:
        with ExcelWriter(file_name, engine='openpyxl', mode='a') as writer: 
            data.to_excel(writer, sheet_name=f'{region_id}', index=False)

    if region_id == '70':
        format_GKV(file_name)
        format_GKV_104(file_name)
        send_email(mail, text, msg_text=name, files=[file_name])
        os.replace(file_name, f'{trek}/backup/{new_file_name}') 

    writing_to_log_file(name_log, text)

# *****************************************************************
def generating_report_GKV(df, name_log, name, mail):
    file_name = f'{name}.xlsx'
    text = f'{name} на {today}'
    new_file_name = f'{today} - {file_name}'

    data = pd.DataFrame(df)
    data.to_excel(file_name, index=False)
    # format_GKV(file_name)

    send_email(mail, text, msg_text=text, files=[file_name])
    os.replace(file_name, f'{trek}/backup/{new_file_name}') 
    writing_to_log_file(name_log, text)      

# *****************************************************************
def format_GKV_104(file):
    wb = load_workbook(file)
    sheet = wb['104']
    list_formula = ['C', 'D', 'E']
    list_not = [13, 14, 15, 18, 19, 20, 25, 26, 36, 37, 44, 45, 49, 50, 52, 53]
    for ceil in list_formula:
        for i in range(3, 54):
            if list_not.count(i) == 0:
                sheet[f'{ceil}{i}'] = f"=SUM('58:70'!{ceil}{i})"
    wb.save(file)

# *****************************************************************
def format_GKV(file):
    # определим стили сторон
    style_g = 'thin'
    b_left=Side(border_style=style_g, color='FF000000')
    b_right=Side(border_style=style_g, color='FF000000')
    b_top=Side(border_style=style_g, color='FF000000')
    b_bottom=Side(border_style=style_g, color='FF000000')
    b_diagonal=Side(border_style=None, color='FF000000')
    b_diagonal_direction=0
    b_outline=Side(border_style=None, color='FF000000')
    b_vertical=Side(border_style=None, color='FF000000')
    b_horizontal=Side(border_style=None, color='FF000000')

    wb = load_workbook(file)

    for sh in wb.sheetnames:
        sheet = wb[sh]
        for row in sheet.iter_rows():
            for cell in row:
        #         выравнивание текста во всех ячейках
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        #         рисуем границы
                cell.border = Border(top=b_top, bottom=b_bottom, left=b_left, right=b_right)

        for i in range(1, 14):
            sheet[f'B{i}'].alignment = Alignment(wrap_text=True, horizontal='left')

        list_heading = ['B1', 'B2', 'B15', 'B19', 'B26', 'B37', 'B45', 'B50']
        for ceil in list_heading: # выравнивание текста в определенных ячейках
            sheet[ceil].alignment = Alignment(horizontal='center', vertical='center')
            sheet[ceil].font = Font(bold=True)

        list_itog = ['B14', 'B18', 'B25', 'B36', 'B44', 'B49', 'B52', 'B53']
        for ceil in list_itog: # выравнивание текста в определенных ячейках
            sheet[ceil].alignment = Alignment(horizontal='right', vertical='center')

        list_rovno = ['B2:E2', 'B15:E15', 'B19:E19', 'B26:E26', 'B37:E37', 'B45:E45', 'B50:E50']
        for ceil in list_rovno: # объедение ячеек
            sheet.merge_cells(ceil)

        width_stolb = {'B' : 120, 'C' : 40, 'D' : 20, 'E' : 20}
        for key, value in width_stolb.items(): # ширина столбцов
            sheet.column_dimensions[key].width = value

        list_formula = ['C', 'D', 'E']
        for ceil in list_formula:
            sheet[f'{ceil}14'] = f'=SUM({ceil}3:{ceil}12)'
            sheet[f'{ceil}13'] = f'=SUM({ceil}9:{ceil}11)'
            sheet[f'{ceil}18'] = f'=SUM({ceil}16:{ceil}17)'
            sheet[f'{ceil}20'] = f'=SUM({ceil}21:{ceil}23)'
            sheet[f'{ceil}25'] = f'=SUM({ceil}21:{ceil}24)'
            sheet[f'{ceil}36'] = f'=SUM({ceil}27:{ceil}35)'
            sheet[f'{ceil}44'] = f'=SUM({ceil}38:{ceil}43)'
            sheet[f'{ceil}49'] = f'=SUM({ceil}46:{ceil}48)'
            sheet[f'{ceil}52'] = f'={ceil}51'
            sheet[f'{ceil}53'] = f'={ceil}14+{ceil}18+{ceil}25+{ceil}36+{ceil}44+{ceil}49+{ceil}52'
        
    wb.save(file)

# *****************************************************************
def generating_mail_mo(region_id, mail, test):
    mail_mo = {
        58 : 'uszn@nur.yanao.ru',
        59 : 'szn@krasnoselkupsky.yanao.ru',
        60 : 'dtszns@slh.yanao.ru',
        61 : 'utszn@priuralye.yanao.ru',
        62 : 'mail@utszn.lbt.yanao.ru',
        63 : 'usp@nadym.yanao.ru',
        64 : 'utszn@gubadm.ru',
        65 : 'uszn@muravlenko.yanao.ru',
        66 : 'ORSiDP@noyabrsk.yanao.ru',
        67 : 'usp@pur.yanao.ru',
        68 : 'uszn@yam.yanao.ru',
        69 : 'uszn@shur.yanao.ru',
        70 : 'sz@tazovsky.yanao.ru',
    }
    
    if test == 1:
        return mail
    else:
        mail = f'{mail}, {mail_mo[region_id]}'
        return mail
# *****************************************************************
def generating_report_GSP(df, name_log, name, mail):
    file_name = f'{name}.xlsx'
    data = pd.DataFrame(df)
    data.to_excel(file_name, index=False)

    text = f'{file_name} на {today}'

    send_email(mail, text, msg_text=file_name, files=[file_name])

    new_file_name = f'{today} - {file_name}'
    os.replace(file_name, f'{trek}/backup/{new_file_name}') 

    writing_to_log_file(name_log, text)

# *****************************************************************
def send_email(addr_to, msg_subj, msg_text='', files=[]):
    name_log = 'smtp'
    writing_to_log_file(name_log, '*********************************************')
    writing_to_log_file(name_log, 'Вызов функции send_email')
    writing_to_log_file(name_log, f'Отправка письма {addr_to} с темой "{msg_subj}"')
    # '''Функция по отпрвке писем через smtp.yanao.ru. 
    #    Логин, пароль и сервер указываются в файле access_mail.txt.
    #    addr_to - указать адресата (обязательно).
    #    msg_subj - указать тему (обязательно).
    #    msg_text - указать текст письма (необязательно).
    #    files - указать файлы в виде ['файл 1', 'файл 2', и т.д.] (необязательно).
    # '''
    path = f'{trek}/access_mail.txt'
    with open(path) as f:
        access = json.load(f)
    
    server = access['server'] # сервер отправки
    addr_from = access['user'] # Отправитель
    password  = access['passwd'] # Пароль

    msg = MIMEMultipart() # Создаем сообщение
    msg['From']    = addr_from # Адресат
    msg['To']      = addr_to # Получатель
    msg['Subject'] = msg_subj # Тема сообщения

    body = msg_text # Текст сообщения
    msg.attach(MIMEText(body, 'plain')) # Добавляем в сообщение текст

    process_attachement(msg, files)

    try:
        # подключаемся к почтовому сервису
        smtp = smtplib.SMTP(server) # Создаем объект SMTP
        writing_to_log_file(name_log, f'Подлючение к серверу {server}')
#         smtp.set_debuglevel(1) # журнал, при необходимости включаем
#         smtp.set_debuglevel(True) # Включаем режим отладки, если не нужен - можно закомментировать
        smtp.starttls() # Начинаем шифрованный обмен по TLS
        smtp.ehlo()
        # логинимся на почтовом сервере
        smtp.login(addr_from, password) # Получаем доступ
        smtp.send_message(msg) # Отправляем сообщение
        writing_to_log_file(name_log, 'Письмо отправлено')
    except smtplib.SMTPException as err:
        writing_to_log_file(name_log, f'Ошибка подлючения к серверу {server} - {err}')
        # print('Что - то пошло не так...')
        raise err
    finally:
        smtp.quit()


def process_attachement(msg, files):                        # Функция по обработке списка, добавляемых к сообщению файлов
    name_log = 'smtp'
    writing_to_log_file(name_log, 'Вызов функции process_attachement')
    # '''Функция по обработке списка, добавляемых к сообщению файлов'''
    for f in files:
        if os.path.isfile(f):                               # Если файл существует
            attach_file(msg, f)                             # Добавляем файл к сообщению
        elif os.path.exists(f):                             # Если путь не файл и существует, значит - папка
            dir = os.listdir(f)                             # Получаем список файлов в папке
            for file in dir:                                # Перебираем все файлы и...
                attach_file(msg, f+"/"+file)                # ...добавляем каждый файл к сообщению



def attach_file(msg, filepath):                             # Функция по добавлению конкретного файла к сообщению
    name_log = 'smtp'
    writing_to_log_file(name_log, 'Вызов функции attach_file')
    # '''Функция по добавлению конкретного файла к сообщению'''
    filename = os.path.basename(filepath)                   # Получаем только имя файла
    ctype, encoding = mimetypes.guess_type(filepath)        # Определяем тип файла на основе его расширения
    
    if ctype is None or encoding is not None:               # Если тип файла не определяется
        ctype = 'application/octet-stream'                  # Будем использовать общий тип
    maintype, subtype = ctype.split('/', 1)                 # Получаем тип и подтип
    
    if maintype == 'text':                                  # Если текстовый файл
        with open(filepath) as fp:                          # Открываем файл для чтения
            file = MIMEText(fp.read(), _subtype=subtype)    # Используем тип MIMEText
            fp.close()                                      # После использования файл обязательно нужно закрыть
    elif maintype == 'image':                               # Если изображение
        with open(filepath, 'rb') as fp:
            file = MIMEImage(fp.read(), _subtype=subtype)
            fp.close()
    elif maintype == 'audio':                               # Если аудио
        with open(filepath, 'rb') as fp:
            file = MIMEAudio(fp.read(), _subtype=subtype)
            fp.close()
    else:                                                   # Неизвестный тип файла
        with open(filepath, 'rb') as fp:
            file = MIMEBase(maintype, subtype)              # Используем общий MIME-тип
            file.set_payload(fp.read())                     # Добавляем содержимое общего типа (полезную нагрузку)
            fp.close()
            encoders.encode_base64(file)                    # Содержимое должно кодироваться как Base64
    file.add_header('Content-Disposition', 'attachment', filename=filename) # Добавляем заголовки
    msg.attach(file)                                        # Присоединяем файл к сообщению

# *****************************************************************
def writing_to_log_file(file, text):
    # '''функция создания лог файла и записи в него информации'''
    dt = datetime.now().strftime('%Y-%m-%d %X')
    path = f'{trek}/log/{file}.log'
    # with open(path, 'a+', encoding='utf8') as file_log:
    with open(path, 'a+') as file_log:
        log = f'{dt} : {text} \n'
        file_log.write(log)

# *****************************************************************
def connect_oracle():
    mail = f'IVAbdulganiev@yanao.ru, 300195@mail.ru'
    name_log = 'access_report'

    #***************************************************************
    path = f'{trek}/access_report.txt'
    with open(path) as f:
        access = json.load(f)
    
    driver = f'{trek}/ojdbc14.jar'
    path_base = access['path_base']
    password = access['password']
    login = access['login']
    port = access['port']
    sid = access['sid']

    try:
        conn = jaydebeapi.connect(
            'oracle.jdbc.driver.OracleDriver',
            f'jdbc:oracle:thin:{login}/{password}@{path_base}:{port}/{sid}',
            [login, password],
            driver)
    except Exception as e:
        text = f'произошла ошибка при вызове функции connect_oracle - {e}'
        alarm_log(mail, name_log, text)

    curs = conn.cursor()

    return curs

# *****************************************************************
def connect_oracle_large_family():
    mail = f'IVAbdulganiev@yanao.ru, 300195@mail.ru'
    name_log = 'access_report_large_family'

    #***************************************************************
    path = f'{trek}/access_large_family.txt'
    with open(path) as f:
        access = json.load(f)
    
    driver = f'{trek}/ojdbc14.jar'
    path_base = access['path_base']
    password = access['password']
    login = access['login']
    port = access['port']
    sid = access['sid']

    try:
        conn = jaydebeapi.connect(
            'oracle.jdbc.driver.OracleDriver',
            f'jdbc:oracle:thin:{login}/{password}@{path_base}:{port}/{sid}',
            [login, password],
            driver)
    except Exception as e:
        text = f'произошла ошибка при вызове функции connect_oracle_large_family - {e}'
        alarm_log(mail, name_log, text)

    curs = conn.cursor()

    return curs

# *****************************************************************
def connect_oracle_public_service():
    mail = f'IVAbdulganiev@yanao.ru, 300195@mail.ru'
    name_log = 'access_report_public_service'

    #***************************************************************
    path = f'{trek}/access_public_service.txt'
    with open(path) as f:
        access = json.load(f)
    
    driver = f'{trek}/ojdbc14.jar'
    path_base = access['path_base']
    password = access['password']
    login = access['login']
    port = access['port']
    sid = access['sid']

    try:
        conn = jaydebeapi.connect(
            'oracle.jdbc.driver.OracleDriver',
            f'jdbc:oracle:thin:{login}/{password}@{path_base}:{port}/{sid}',
            [login, password],
            driver)
    except Exception as e:
        text = f'произошла ошибка при вызове функции connect_oracle_public_service - {e}'
        alarm_log(mail, name_log, text)

    curs = conn.cursor()

    return curs

# *****************************************************************
def connect_oracle_ecert():
    mail = f'IVAbdulganiev@yanao.ru, 300195@mail.ru'
    name_log = 'connect_oracle_ecert'

    #***************************************************************
    path = f'{trek}/access_ecert.txt'
    with open(path) as f:
        access = json.load(f)
    
    driver = f'{trek}/ojdbc14.jar'
    path_base = access['path_base']
    password = access['password']
    login = access['login']
    port = access['port']
    sid = access['sid']

    try:
        conn = jaydebeapi.connect(
            'oracle.jdbc.driver.OracleDriver',
            f'jdbc:oracle:thin:{login}/{password}@{path_base}:{port}/{sid}',
            [login, password],
            driver)
    except Exception as e:
        text = f'произошла ошибка при вызове функции connect_oracle_ecert - {e}'
        alarm_log(mail, name_log, text)

    curs = conn.cursor()

    return curs

# *****************************************************************
def movi_vipnet(test, file, name_log, name_def, recipient='uszn'):
    os.chmod(file, 0o666)
    t = os.getcwd()
    copy_vipnet(test, file, name_log, name_def, recipient)
    try:
        os.remove(file)
        writing_to_log_file(name_log, f'{file} удален в папке {t}')
    except Exception as e:
        text = f'os.remove - {e} - в папке {t}'
        writing_to_log_file(name_log, text)    

# *****************************************************************
def copy_vipnet(test, file, name_log, name_def, recipient='uszn'):
    os.chmod(file, 0o666)
    t = os.getcwd()
    recip = {
        'uszn' : [vipnet_test, vipnet_main],
        'cso' : [vipnet_test_cso, vipnet_main_cso],
        'do' : [vipnet_test_do, vipnet_main_do],
        }
    if test == 1:
        path = recip[recipient][0]
    else:
        path = recip[recipient][1]
    try:
      shutil.copyfile(file, f'{path}/{file}')
    except Exception as e:
      send_email(mail, f'{name_def} - ошибка копирования файла', msg_text=file)
      text = f'{name_def} - ошибка копирования файла {file} из папки {t} в папку {path} - {e}'
      writing_to_log_file(name_log, text)

#***************************************************************
def backup_file(test, file, name_log, name_def, path_backup = f'{trek}/backup/', path_out = f'{trek}/'):
    os.chmod(f'{path_out}/{file}', 0o666)
    new_file_name = f'{today}_{sec} - {file}'
    if test == 1:
        path_backup = f'{trek}/backup1/'
    try:
        shutil.move(f'{path_out}/{file}', f'{path_backup}/{new_file_name}')
        writing_to_log_file(name_log, f'Файл {file} из папки {path_out} перемещен в {path_backup} и переименован в {new_file_name}')
    except Exception as e:
        send_email(mail, f'{name_def} - ошибка переноса файла', msg_text=file)
        text = f'{name_def} - ошибка переноса файла {file} из папки {path_out} в папку {path_backup} - {e}'
        writing_to_log_file(name_log, text)      

#***************************************************************
def backup_file_pfr_4454(test, file, name_log, name_def, path):
    os.chmod(file, 0o666)
    new_file_name = f'{today} - {file}'
    if test == 1:
        path_backup = f'{trek}/backup1/'
    else:
        path_backup = f'{trek}/backup/'
    try:
        shutil.move(f'{path}/{file}', f'{path_backup}/{new_file_name}')
        writing_to_log_file(name_log, f'Файл {file} перемещен в {path_backup} и переименован в {new_file_name}')
    except Exception as e:
        send_email(mail, f'{name_def} - ошибка переноса файла', msg_text=file)
        text = f'{name_def} - ошибка переноса файла {file} в папку {path_backup} - {e}'
        writing_to_log_file(name_log, text)      

# *****************************************************************
def generating_report_files(df, name_log, name_def, test, mail, recipient='uszn'):
    data = pd.DataFrame(df)
    files = ''
    mo = set(data['name'])
    if len(mo) > 0:
        for row in mo:
            file = row + '.xlsx'
            data[data['name'].isin([row])].to_excel(file, index=False)
            movi_vipnet(test, file, name_log, name_def, recipient)
            files += file + '\n'
        writing_to_log_file(name_log, '\n'+files)
        send_email(mail, f'{name_def} отправлены', msg_text=files)
    else:
        text = 'файлов нет'
        writing_to_log_file(name_log, text)


# *****************************************************************
def generating_report_files_PFR(df, name_log, name_def, test, mail):
    data = pd.DataFrame(df)
    dt = datetime.now().strftime('%m-%Y')
    file = f'{name_def}{dt}.xlsx'
    data.to_excel(file, index=False)
    writing_to_log_file(name_log, f'создали файл {file}')

    try:
        path_backup = f'{trek}/backup/na_3/'
        writing_to_log_file(name_log, f'path_backup {path_backup}')
        
        try:
            new_file = file + '.zip'
            writing_to_log_file(name_log, f'архирование файла {new_file}')
            zipFile = zipfile.ZipFile(new_file, 'w', zipfile.ZIP_DEFLATED)
            zipFile.write(file)
            zipFile.close()
        except:
            writing_to_log_file(name_log, 'ошибка архивирования файла {file} в {new_file}')
        
        movi_vipnet(test, new_file, name_log, name_def)

        backup_file(test, file, name_log, name_def, path_backup)

        writing_to_log_file(name_log, new_file)

        send_email(mail, f'{name_def} в ПФР отправлен', msg_text=new_file)
        writing_to_log_file(name_log, f'{name_def} в ПФР отправлен')
    except Exception as e:
        text = f'{name_def} - ошибка'
        send_email(mail, f'{name_def} - ошибка ', msg_text=text)
        writing_to_log_file(name_log, text)      

# *****************************************************************
def generating_report_files_PFR_2(name_log, name_def, test, mail, text):
    dt = datetime.now().strftime('%m-%Y')
    new_file = name_def
    path_backup = f'{trek}/backup/RSD/'

    try:
        os.remove(f'{path}{new_file}')
        writing_to_log_file(name_log, f'{path}{new_file} удален')
    except Exception as e:
        writing_to_log_file(name_log, f'сбой при удалении {path}{new_file}, ошибка {e}')

    try:
        os.remove(f'{path_backup}{new_file}')
        writing_to_log_file(name_log, f'{path_backup}{new_file} удален')
    except Exception as e:
        writing_to_log_file(name_log, f'сбой при удалении {path_backup}{new_file}, ошибка {e}')
    
    cnt = 0 # счетчик для отправки

    movi_vipnet(test, new_file, name_log, name_def)
    cnt += 1
    backup_file(test, new_file, name_log, name_def, path_backup)
    cnt += 1

    if cnt == 2:
        send_email(mail, f'{name_def} в ПФР отправлен', msg_text=text)
        writing_to_log_file(name_log, f'Письмо отправлено с текстом - {text}')
    else:
        send_email(mail, f'Alarm {name_def}', msg_text=text)
        writing_to_log_file(name_log, 'Письмо "Alarm" отправлено')

# ***************************************************************************************
def count_line_table(curs, table, name_log):
    cnt = 0
    try:
        curs.execute(f'SELECT COUNT(1) FROM {table}')
        cnt = curs.fetchall()[0][0]
        text = f' в {table} кол-во строк {cnt}'
        writing_to_log_file(name_log, text)
        check = 0
    except:
        text = f'{table} нету'
        writing_to_log_file(name_log, text)
    return cnt

# ***************************************************************************************
def grant_table(curs, table, role, name_log):
    check = 0

    try:
        curs.execute(f'grant select on {table} to {role}')
        text = f'grant select on {table} to {role}'
        writing_to_log_file(name_log, text)
        check = 0
    except Exception as e:
        text = f'{role} - alarm - {alarm}'
        writing_to_log_file(name_log, text)
        send_email(mail, f'Alarm - {name_log}', msg_text=text)
        check = 0

    return check

# ***************************************************************************************
def report_temp_table(name_log, mail, table, file_sql):
    writing_to_log_file(name_log, '**********start**************')
    curs = connect_oracle()
    check = count_line_table(curs, table, name_log)

    try:
        curs.execute(f'DROP TABLE {table}')
        text = f'{table} удалена'
        writing_to_log_file(name_log, text)
    except:
        text = f'{table} нету'
        writing_to_log_file(name_log, text)

    try:
        text = f'обновление {table}'
        writing_to_log_file(name_log, text)
        
        with open(file_sql, 'r', encoding='utf8') as f:
            sql = f.read()
        curs.execute(sql)
        
        text = f'{table} обновлена'
        writing_to_log_file(name_log, text)
        check = 0
    except Exception as e:
        alarm = str(e)
        text = f'alarm - {alarm} - {sql}'
        writing_to_log_file(name_log, text)
        send_email(mail, f'Alarm - {name_log}', msg_text=text)
        check = 1
        
    if check == 0:
        check = grant_table(curs, table, 'USZN_TSRV_ROLE_000000000003', name_log)
    if check == 0:
        check = grant_table(curs, table, 'USZN_TSRV_ROLE_000000000004', name_log)
    if check == 0:
        check = grant_table(curs, table, 'USZN_TSRV_ROLE_000000000008', name_log)
    if check == 0:
        check = grant_table(curs, table, 'USZN_TSRV_ROLE_000000000010', name_log)

    cnt = count_line_table(curs, table, name_log)
    text = f' в {table} кол-во строк {cnt}'
    writing_to_log_file(name_log, text)
    send_email(mail, name_log, msg_text=text)

    writing_to_log_file(name_log, '**********start**************')

#***************************************************************
def insert_temp_table(name_log, mail, table, file_sql):
    writing_to_log_file(name_log, '**********start**************')
    curs = connect_oracle()

    try:
        curs.execute(f'DELETE FROM {table}')
        text = f'{table} очищена'
        writing_to_log_file(name_log, text)
        check = 0
    except:
        text = f'{table} нету'
        writing_to_log_file(name_log, text)
        check = 1

    if check == 0:
        try:
            text = f'обновление {table}'
            writing_to_log_file(name_log, text)
            
            with open(file_sql, 'r', encoding='utf8') as f:
                sql = f.read()
            curs.execute(sql)
            
            text = f'{table} обновлена'
            writing_to_log_file(name_log, text)
            check = 0
        except Exception as e:
            alarm = str(e)
            text = f'alarm - {alarm} - {sql}'
            writing_to_log_file(name_log, text)
            send_email(mail, f'Alarm - {name_log}', msg_text=text)
            check = 1
    
    if check == 0:
        check = grant_table(curs, table, 'USZN_TSRV_ROLE_000000000003', name_log)
    if check == 0:
        check = grant_table(curs, table, 'USZN_TSRV_ROLE_000000000004', name_log)
    if check == 0:
        check = grant_table(curs, table, 'USZN_TSRV_ROLE_000000000008', name_log)
    if check == 0:
        check = grant_table(curs, table, 'USZN_TSRV_ROLE_000000000010', name_log)

    cnt = count_line_table(curs, table, name_log)
    text = f'в {table} кол-во строк {cnt}'
    writing_to_log_file(name_log, text)
    send_email(mail, name_log, msg_text=text)

    writing_to_log_file(name_log, '**********start**************')

#***************************************************************
def write_file(file, log):
    try:
        xl = pd.read_excel(file)
        writing_to_log_file(log, f'Файл {file} записан в dataframe')
        return xl
    except Exception as e:
        writing_to_log_file(log, f'Alarm: \n {e}')

#***************************************************************
def write_file_object(file, log):
    try:
        xl = pd.read_excel(file, dtype="object")
        writing_to_log_file(log, f'Файл {file} записан в dataframe')
        return xl
    except Exception as e:
        writing_to_log_file(log, f'Alarm: \n {e}')

#***************************************************************
def dat(x):
    try:
        x = str(x).replace('\n', '')
        x = str(x).replace(' 00:00:00', '')
        return re.sub(r'(\d{4})-(\d{2})-(\d{2})', r'\3.\2.\1', x)
    except:
        return x
    
#***************************************************************
def up(x):
    try:
        return x.upper()
    except:
        return x
    
#***************************************************************
def mo_id(x):
    x = x.upper()
    x = x.replace('МО ', '')
    x = x.replace('Р-Н', '')
    x = x.replace('РАЙОН', '')
    x = x.replace('Г.', '')
    x = x.replace(' ', '')
    x = x.replace('-', '')
    
    try:
        mo = {
        'НОВЫЙУРЕНГОЙ' : 58,
        'НОВЫЙ-УРЕНГОЙ' : 58,
        'КРАСНОСЕЛЬКУПСКИЙ' : 59,
        'САЛЕХАРД' : 60,
        'ПРИУРАЛЬСКИЙ' : 61,
        'ЛАБЫТНАНГИ' : 62,
        'НАДЫМСКИЙ' : 63,
        'НАДЫМ' : 63,
        'ГУБКИНСКИЙ' : 64,
        'МУРАВЛЕНКО' : 65,
        'НОЯБРЬСК' : 66,
        'ПУРОВСКИЙ' : 67,
        'ЯМАЛЬСКИЙ' : 68,
        'ШУРЫШКАРСКИЙ' : 69,
        'ТАЗОВСКИЙ' : 70,
        }
        return mo[x]
    except:
        return 104

# *************************************************
def num_to_str(x):
    try:
        return str(int(x))
    except:
        return x

# *************************************************
def staples(x):
    try:
        return x.replace('[', '').replace(']', '')
    except:
        return x

# *************************************************
def num_int_to_str(x):
    try:
        return str(x)
    except:
        return x

# *************************************************
def snils(x):
    try:
        x = num_to_str(x)
        if len(x) == 10:
            x = '0'+ x
        elif len(x) == 9:
            x = '00'+ x
        return re.sub(r'(\d{3})(\d{3})(\d{3})(\d{2})', r'\1-\2-\3 \4', x)
    except:
        return x

# *************************************************
def replace_nat(x):
    try:
        return x.replace('NaT', '')
    except:
        return x

# *************************************************
def replace_nan(x):
    try:
        return x.replace('nan', '')
        return x.replace('None', '')
    except:
        return x

# *************************************************    
def sfr_uszn_kod_name(x, i):
    mo = {
   2575.000002 : [58, 'Управление по труду и социальной защите населения Администрации города Новый Уренгой'],
   2575.000159 : [58, 'Департамент социальной политики Администрации города Новый Уренгой'],
   2575.000003 : [59, 'Управление по труду и социальной защите населения Администрации Красноселькупского района'],
   2575.000004 : [60, 'Департамент по труду и социальной защите населения Администрации муниципального образования город Салехард'],
   2575.000048 : [61, 'Управление по труду и социальной защите населения Администрации Приуральского района'],
   2575.000151 : [62, 'Муниципальное учреждение Управление по труду и социальной защите населения Администрации города Лабытнанги'],
   2575.000005 : [63, 'Управление социальных программ Администрации Надымского район'],
   2575.000006 : [64, 'Муниципальное учреждение “Управление по труду и социальной защите населения Администрации города Губкинского”'],
   2575.000141 : [65, 'Управление социальной защиты населения Администрации города Муравленко'],
   2575.000108 : [66, 'Управление социальной защиты населения Администрации города Ноябрьска'],
   2575.000125 : [67, 'Управление социальной политики администрации Пуровского района'],
   2575.000080 : [68, 'Департамент социальной защиты населения Администрации Ямальский район'],
   2575.000081 : [69, 'Департамент социальной защиты населения Администрации муниципального образования Шурышкарский район'],
   2575.000082 : [70, 'Департамент социального развития Администрации Тазовского района'],
   2575.000001 : [71, 'Департамент социальной защиты населения Ямало-Ненецкого автономного округа']
    }
    if i == 0:
        return mo[x][i]
    elif i == 1:
        return mo[x][i]
    
# *************************************************
def kod_sfr_uszn_id(x):
    return sfr_uszn_kod_name(x, 0)

# *************************************************
def kod_sfr_uszn_name(x):
    return sfr_uszn_kod_name(x, 1)

# *****************************************************************
def copy_file(file, name_log, name_def, path_in, path_to):
    t = os.getcwd()
    writing_to_log_file(name_log, f'каталог {t} с файлом {file}')
    try:
        os.chmod(f'{path_in}{file}', 0o666)
        writing_to_log_file(name_log, 'chmod')
    except Exception as e:
        text = f'os.chmod - {e} - в папке {path_in}'
        writing_to_log_file(name_log, text)

    try:
        shutil.copyfile(f'{path_in}{file}', f'{path_to}{file}')
        writing_to_log_file(name_log, f'{path_in}{file} - {path_to}{file}')
    except Exception as e:
        send_email(mail, f'{name_def} - ошибка копирования файла', msg_text=file)
        text = f'{name_def} - {e} - ошибка копирования файла {path_in}{file} в папку {path_to}'
        writing_to_log_file(name_log, text)

# *****************************************************************
def movi_file(file, name_log, name_def, path_in, path_to):
    t = os.getcwd()
    writing_to_log_file(name_log, f'каталог {t} с файлом {file}')
    try:
        os.chmod(f'{path_in}{file}', 0o666)
        writing_to_log_file(name_log, 'chmod')
    except Exception as e:
        text = f'os.chmod - {e} - в папке {path_in}'
        writing_to_log_file(name_log, text)
    copy_file(file, name_log, name_def, path_in, path_to)

    try:
        os.remove(file)
        writing_to_log_file(name_log, f'{file} удален в папке {path_in}')
    except Exception as e:
        text = f'os.remove - {e} - в папке {path_in}'
        writing_to_log_file(name_log, text)
